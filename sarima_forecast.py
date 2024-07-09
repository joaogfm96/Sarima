import pandas as pd
from pmdarima import auto_arima
from openpyxl import load_workbook
from datetime import timedelta

# Load data from the Excel spreadsheet
file_path = r'C:\Users\Joao Mosqueira\Documents\Previsão - Fluxo de Caixa\CashFlow.xlsm'
sheet_name_consolidado = 'Consolidado'
sheet_name_forecast_sarima = 'Forecast - SARIMA (3)'

# Read the Consolidado sheet
df = pd.read_excel(file_path, sheet_name=sheet_name_consolidado, header=[2])

# Identify the last date in the dataset
ultima_data = pd.to_datetime(df['Data'].iloc[-1])

# Number of business days to forecast from cell A2 in the Forecast - SARIMA (3) sheet
wb = load_workbook(file_path, keep_vba=True)
ws_forecast_sarima = wb.create_sheet(sheet_name_forecast_sarima) if sheet_name_forecast_sarima not in wb.sheetnames else wb[sheet_name_forecast_sarima]
dias_para_prever = wb[sheet_name_forecast_sarima]['A2'].value

# Check if dias_para_prever is being read correctly
print(f"Days to forecast: {dias_para_prever}")

# Number of historical days to be used
dias_historicos = dias_para_prever * 3

# Initialize DataFrame for forecasts
forecasts = pd.DataFrame()
print(forecasts)

# Add the date column to the forecasts
forecasts['Data'] = pd.date_range(start=ultima_data + timedelta(days=1), periods=dias_para_prever, freq='B')
print(forecasts)

# Iterate over each column (excluding 'Data')
for coluna in df.columns[1:]:
    # Select the last dias_historicos business days of history
    temp_df = df[['Data', coluna]].tail(dias_historicos).rename(columns={'Data': 'ds', coluna: 'y'})
    temp_df['y'] = temp_df['y'].replace('[\$,]', '', regex=True).astype(float)
    
    # Replace zero and NaN values with 0.000000001
    temp_df['y'] = temp_df['y'].replace(0, 0.000000001).fillna(0.000000001)
    
    # Check the content of temp_df
    print(f"Checking column: {coluna}")
    print(temp_df.tail())

    # Check if there are at least 2 non-null rows
    if temp_df.dropna().shape[0] < 2:
        print(f"Column {coluna} does not have enough data for forecasting.")
        continue
    
    # Instantiate and fit the SARIMA model automatically
    try:
        model = auto_arima(temp_df['y'], seasonal=True, m=12, suppress_warnings=True)
        
        # Print the best parameters found
        print(f"Best parameters for column {coluna}: {model.order} with seasonality {model.seasonal_order}")
        
        # Make future forecasts
        forecast = model.predict(n_periods=dias_para_prever)
        print(f"Forecast for column {coluna}: {forecast}")
        
        # Add forecasts to the forecasts DataFrame
        forecasts[coluna] = forecast
    except Exception as e:
        print(f"Error fitting SARIMA model for column {coluna}: {e}")
        continue

# Check forecasts before writing to Excel
print("Final forecasts:")
print(forecasts)

# Write forecasts back to the Forecast - SARIMA (3) sheet
# Write the date column
for idx, date in enumerate(forecasts['Data'], start=4):
    ws_forecast_sarima.cell(row=idx, column=1, value=date)

# Write the forecasts
for col_idx, coluna in enumerate(forecasts.columns[1:], start=2):
    for row_idx, value in enumerate(forecasts[coluna], start=4):
        ws_forecast_sarima.cell(row=row_idx, column=col_idx, value=value)

# Save the Excel file with the forecasts
wb.save(file_path)

print("Forecasts added to the Excel file.")
